import json
import re
from datetime import datetime, UTC
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parents[1]
RAW = BASE / "data" / "raw"
OUT = BASE / "data" / "processed" / "electricity_dashboard_data.json"
OUT_JS = BASE / "data" / "processed" / "electricity_dashboard_data.js"

SOURCE_URLS = {
    "Electricity_since_1920.xlsx": "https://assets.publishing.service.gov.uk/media/6889f86f76f68cc8414d5b6d/Electricity_since_1920.xlsx",
    "DUKES_5.1.xlsx": "https://assets.publishing.service.gov.uk/media/688a2873a11f859994409277/DUKES_5.1.xlsx",
    "DUKES_5.3.xlsx": "https://assets.publishing.service.gov.uk/media/688a288b6478525675739052/DUKES_5.3.xlsx",
    "DUKES_5.5.xlsx": "https://assets.publishing.service.gov.uk/media/688a28a576f68cc8414d5bd0/DUKES_5.5.xlsx",
    "DUKES_5.7.xlsx": "https://assets.publishing.service.gov.uk/media/688a28b38b3a37b63e739064/DUKES_5.7.xlsx",
    "DUKES_5.14.xlsx": "https://assets.publishing.service.gov.uk/media/688a2923a11f85999440927c/DUKES_5.14.xlsx",
    "ET_5.1_DEC_25.xlsx": "https://assets.publishing.service.gov.uk/media/6942c218143d960161547dd9/ET_5.1_DEC_25.xlsx",
    "ET_5.2_DEC_25.xlsx": "https://assets.publishing.service.gov.uk/media/6942c222fdbd8404f9e1f24a/ET_5.2_DEC_25.xlsx",
    "ET_5.3_FEB_26.xlsx": "https://assets.publishing.service.gov.uk/media/699d7ff7532c9ad91ebbcbb7/ET_5.3_FEB_26.xlsx",
    "ET_5.6_DEC_25.xlsx": "https://assets.publishing.service.gov.uk/media/6942c24c9273c48f554cf530/ET_5.6_DEC_25.xlsx",
}


def clean_label(value):
    if value is None:
        return ""
    label = str(value).replace("\n", " ").strip()
    label = re.sub(r"\s*\[note[^\]]*\]", "", label, flags=re.IGNORECASE)
    label = re.sub(r"\s+", " ", label)
    return label.strip()


def is_year(v):
    if isinstance(v, int):
        return 1800 <= v <= 2100
    if isinstance(v, float) and v.is_integer():
        return 1800 <= int(v) <= 2100
    if isinstance(v, str) and v.strip().isdigit():
        y = int(v.strip())
        return 1800 <= y <= 2100
    return False


def to_number(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def parse_historical_sheet(ws, unit, wanted_metrics):
    header_row = None
    for r in range(1, 40):
        if clean_label(ws.cell(r, 1).value).lower() == "year":
            header_row = r
            break
    if not header_row:
        return None

    headers = {}
    c = 2
    while c <= ws.max_column:
        h = clean_label(ws.cell(header_row, c).value)
        if h:
            headers[c] = h
        c += 1

    series = {h: [] for h in headers.values() if h in wanted_metrics}

    r = header_row + 1
    while r <= ws.max_row:
        year = ws.cell(r, 1).value
        if not is_year(year):
            r += 1
            continue
        year = int(float(year))
        for c, h in headers.items():
            if h not in series:
                continue
            n = to_number(ws.cell(r, c).value)
            if n is not None:
                series[h].append({"x": str(year), "y": n})
        r += 1

    return {
        "unit": unit,
        "series": [{"name": k, "points": v} for k, v in series.items() if v],
    }


def parse_row_labels_year_matrix(ws, header_row, row_start, id_name, title, unit, selected_labels=None, row_label_cols=1):
    year_cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if is_year(v):
            year_cols.append((c, str(int(float(v)))))

    if len(year_cols) < 3:
        return None

    series = {}
    blank_streak = 0
    r = row_start
    while r <= ws.max_row:
        values = [ws.cell(r, c).value for c in range(1, row_label_cols + 1)]
        text0 = clean_label(values[0]) if values else ""

        if text0.lower().startswith("table ") and r > row_start:
            break

        if not any(v is not None and str(v).strip() for v in values):
            blank_streak += 1
            if blank_streak >= 3:
                break
            r += 1
            continue

        blank_streak = 0
        labels = [clean_label(v) for v in values if clean_label(v)]
        label = " | ".join(labels)

        if selected_labels and label not in selected_labels:
            r += 1
            continue

        if label:
            points = []
            for c, x in year_cols:
                n = to_number(ws.cell(r, c).value)
                if n is not None:
                    points.append({"x": x, "y": n})
            if points:
                series[label] = points
        r += 1

    return {
        "id": id_name,
        "title": title,
        "unit": unit,
        "frequency": "annual",
        "series": [{"name": k, "points": v} for k, v in series.items()],
    }


def parse_et53_month(ws):
    header = 6
    metric_cols = {}
    for c in range(2, 14):
        name = clean_label(ws.cell(header, c).value)
        if name:
            metric_cols[c] = name

    rows = []
    for r in range(7, ws.max_row + 1):
        month = ws.cell(r, 1).value
        if not month:
            continue
        month_label = clean_label(month)
        if not re.search(r"\d{4}", month_label):
            continue
        rows.append((r, month_label))

    rows = rows[-48:]
    series = {name: [] for name in metric_cols.values()}
    for r, month_label in rows:
        for c, name in metric_cols.items():
            n = to_number(ws.cell(r, c).value)
            if n is not None:
                series[name].append({"x": month_label, "y": n})

    keep = {
        "Coal",
        "Gas",
        "Nuclear",
        "Wind",
        "Bioenergy",
        "Solar",
        "Fossil fuels",
        "Low carbon",
    }

    out = []
    for name, pts in series.items():
        normalized = clean_label(name)
        if normalized in keep and pts:
            out.append({"name": normalized, "points": pts})

    return {
        "id": "et53_monthly_fuel_mix_recent",
        "title": "Monthly Fuel Use in Electricity Generation (Recent 48 Months)",
        "unit": "million tonnes of oil equivalent (Mtoe)",
        "frequency": "monthly",
        "note": "This is fuel input energy (Mtoe), not electricity output (GWh/TWh).",
        "convertible_to_twh": True,
        "mtoe_to_twh_factor": 11.63,
        "series": out,
    }


def parse_et56a_quarter(ws):
    header = 7
    metric_cols = {}
    for c in range(3, 18):
        h = clean_label(ws.cell(header, c).value)
        if h:
            metric_cols[c] = h

    series = {h: [] for h in metric_cols.values()}
    for r in range(8, ws.max_row + 1):
        year = ws.cell(r, 1).value
        quarter = clean_label(ws.cell(r, 2).value)
        if not is_year(year) or not quarter.lower().startswith("quarter"):
            continue
        x = f"{quarter} {int(float(year))}"
        for c, h in metric_cols.items():
            n = to_number(ws.cell(r, c).value)
            if n is not None:
                series[h].append({"x": x, "y": n})

    focus = {
        "UK total imports",
        "UK total exports",
        "UK total net imports",
        "Net imports (France to UK)",
        "Net imports (Netherlands to UK)",
    }

    out = []
    for name, pts in series.items():
        n = clean_label(name)
        if n in focus and pts:
            out.append({"name": n, "points": pts[-12:]})

    return {
        "id": "et56_quarterly_trade",
        "title": "Quarterly Electricity Trade Flows",
        "unit": "GWh",
        "frequency": "quarterly",
        "series": out,
    }


def parse_et52_main(ws):
    header = 5
    cols = {}
    for c in range(2, 14):
        h = clean_label(ws.cell(header, c).value)
        if h:
            cols[c] = h

    wanted_rows = {
        "Indigenous production",
        "Imports",
        "Exports",
        "Total supply",
        "Total demand",
        "Final consumption",
    }

    series = {row: [] for row in wanted_rows}
    for r in range(6, 40):
        name = clean_label(ws.cell(r, 1).value)
        if name not in wanted_rows:
            continue
        for c, period in cols.items():
            if "per cent" in period.lower() or period == "-":
                continue
            n = to_number(ws.cell(r, c).value)
            if n is not None:
                series[name].append({"x": period, "y": n})

    return {
        "id": "et52_supply_demand_main",
        "title": "Supply and Demand Main Table (Latest Annual + Quarterly)",
        "unit": "GWh",
        "frequency": "mixed",
        "note": "Contains annual totals and quarterly values in the same table; compare like-with-like periods.",
        "series": [{"name": k, "points": v} for k, v in series.items() if v],
    }


def extract_et51_main_gas_twh_2024():
    wb = openpyxl.load_workbook(RAW / "ET_5.1_DEC_25.xlsx", data_only=True, read_only=True)
    ws = wb["Main Table"]
    # Row 10 is Gas (TWh), col 4 is 2024 value in this sheet layout.
    val = ws.cell(10, 4).value
    return float(val) if isinstance(val, (int, float)) else None


def extract_et53_main_gas_mtoe_2024():
    wb = openpyxl.load_workbook(RAW / "ET_5.3_FEB_26.xlsx", data_only=True, read_only=True)
    ws = wb["Main Table"]
    # Row 8 = 2024, col 5 = Gas.
    val = ws.cell(8, 5).value
    return float(val) if isinstance(val, (int, float)) else None


def build():
    datasets = []

    hist_wb = openpyxl.load_workbook(RAW / "Electricity_since_1920.xlsx", data_only=True, read_only=True)

    hist_generation = parse_historical_sheet(
        hist_wb["Estimated Historical Generation"],
        "GWh",
        {
            "Total estimated generation",
            "Coal estimated generation",
            "Natural gas estimated generation",
            "Nuclear estimated generation",
            "Wind, wave, solar and hydro estimated generation",
        },
    )
    datasets.append(
        {
            "id": "historical_generation_1920",
            "title": "Historical Electricity Generation by Fuel (Estimated)",
            "unit": hist_generation["unit"],
            "frequency": "annual",
            "series": hist_generation["series"],
        }
    )

    hist_supply = parse_historical_sheet(
        hist_wb["Supply, Availability & Consumpt"],
        "TWh",
        {
            "Electricity supplied (net)",
            "Electricity available",
            "Total electricity consumed",
            "Domestic consumption",
            "Industrial consumption",
        },
    )
    datasets.append(
        {
            "id": "historical_supply_consumption",
            "title": "Historical Electricity Supply and Consumption",
            "unit": hist_supply["unit"],
            "frequency": "annual",
            "series": hist_supply["series"],
        }
    )

    hist_capacity = parse_historical_sheet(
        hist_wb["Capacity"],
        "MW",
        {
            "GB Installed Capacity (MW)",
            "GB Output Capacity (MW)",
            "UK declared net capacity (MW)",
            "UK transmission entry capacity (MW)",
            "Simultaneous maximum load met (MW)",
        },
    )
    datasets.append(
        {
            "id": "historical_capacity",
            "title": "Historical Electricity Capacity",
            "unit": hist_capacity["unit"],
            "frequency": "annual",
            "series": hist_capacity["series"],
        }
    )

    dukes_514 = openpyxl.load_workbook(RAW / "DUKES_5.14.xlsx", data_only=True, read_only=True)["5.14"]
    emissions = parse_row_labels_year_matrix(
        dukes_514,
        header_row=4,
        row_start=5,
        id_name="dukes_514_emissions",
        title="Carbon Intensity of Electricity Supplied",
        unit="tonnes CO2 per GWh",
        selected_labels={
            "Coal",
            "Gas",
            "All non renewable fuels",
            "All fuels (including nuclear and renewables)",
        },
    )
    datasets.append(emissions)

    dukes_57 = openpyxl.load_workbook(RAW / "DUKES_5.7.xlsx", data_only=True, read_only=True)["5.7"]
    capacity_mix = parse_row_labels_year_matrix(
        dukes_57,
        header_row=7,
        row_start=31,
        id_name="dukes_57_capacity_mix",
        title="Generating Capacity by Fuel (All Generating Companies)",
        unit="MW",
        selected_labels={
            "All generating companies | Total capacity",
            "All generating companies | Coal fired",
            "All generating companies | Gas fired",
            "All generating companies | Nuclear stations",
            "All generating companies | Onshore wind",
            "All generating companies | Offshore wind",
            "All generating companies | Solar",
            "All generating companies | Bioenergy and waste",
        },
        row_label_cols=2,
    )
    datasets.append(capacity_mix)

    dukes_51 = openpyxl.load_workbook(RAW / "DUKES_5.1.xlsx", data_only=True, read_only=True)["5.1"]
    commodity_balance = parse_row_labels_year_matrix(
        dukes_51,
        header_row=5,
        row_start=6,
        id_name="dukes_51_balance",
        title="Electricity Commodity Balance",
        unit="GWh",
        selected_labels={
            "Production",
            "Imports",
            "Exports",
            "Total supply",
            "Total demand",
            "Final consumption",
        },
    )
    datasets.append(commodity_balance)

    et53 = openpyxl.load_workbook(RAW / "ET_5.3_FEB_26.xlsx", data_only=True, read_only=True)
    datasets.append(parse_et53_month(et53["Month"]))

    et56 = openpyxl.load_workbook(RAW / "ET_5.6_DEC_25.xlsx", data_only=True, read_only=True)
    datasets.append(parse_et56a_quarter(et56["5.6A"]))

    et52 = openpyxl.load_workbook(RAW / "ET_5.2_DEC_25.xlsx", data_only=True, read_only=True)
    datasets.append(parse_et52_main(et52["Main Table"]))

    payload = {
        "generated_at": datetime.now(UTC).isoformat(),
        "collection": "https://www.gov.uk/government/collections/electricity-statistics",
        "sources": [
            {"file": name, "url": url}
            for name, url in SOURCE_URLS.items()
            if (RAW / name).exists()
        ],
        "datasets": [d for d in datasets if d and d.get("series")],
    }

    gas_twh_2024 = extract_et51_main_gas_twh_2024()
    gas_mtoe_2024 = extract_et53_main_gas_mtoe_2024()
    gas_mtoe_to_twh_2024 = gas_mtoe_2024 * 11.63 if gas_mtoe_2024 is not None else None

    payload["unit_audit"] = {
        "checks_ran_at": datetime.now(UTC).isoformat(),
        "gas_2024_cross_check": {
            "et51_gas_twh_2024": gas_twh_2024,
            "et53_gas_mtoe_2024": gas_mtoe_2024,
            "et53_gas_mtoe_as_twh_2024": gas_mtoe_to_twh_2024,
            "difference_twh": None
            if gas_twh_2024 is None or gas_mtoe_to_twh_2024 is None
            else gas_twh_2024 - gas_mtoe_to_twh_2024,
            "note": "Expected to be very close; ET 5.1 reports gas in TWh and ET 5.3 in Mtoe.",
        },
        "dataset_units": [
            {"id": d["id"], "unit": d.get("unit"), "frequency": d.get("frequency")}
            for d in payload["datasets"]
        ],
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    OUT_JS.write_text(
        "window.__ELECTRICITY_DASHBOARD_DATA__ = " + json.dumps(payload) + ";",
        encoding="utf-8",
    )
    print(f"Wrote {OUT}")
    print(f"Wrote {OUT_JS}")
    print(f"Datasets: {len(payload['datasets'])}")


if __name__ == "__main__":
    build()
